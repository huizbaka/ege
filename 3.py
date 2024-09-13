import openpyxl

workbook = openpyxl.load_workbook("3.xlsx")
sheet = workbook["Движение товаров"]

target_markets = {'M1', 'M5', 'M6', 'M10', 'M15'}
total_stock = 0

for row in range(2, sheet.max_row + 1):
    current_market = sheet.cell(row=row, column=3).value
    current_item = sheet.cell(row=row, column=4).value
    operation = sheet.cell(row=row, column=6).value
    quantity = sheet.cell(row=row, column=5).value

    if current_market in target_markets and current_item == 23:
        if operation == "Поступление" and quantity:
            total_stock += quantity
        elif operation == "Продажа" and quantity:
            total_stock -= quantity

print(total_stock)