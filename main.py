import openpyxl

df1 = openpyxl.load_workbook("3.xlsx")
wb = df1["Движение товаров"]

needed_markets = []

wb1 = df1["Магазин"]
for i in range(2, wb1.max_row + 1):
    block = wb1[f"B{i}"].value
    if block == "Октябрьский":
        needed_markets.append(wb1[f"A{i}"].value)

in_stock = 0

for i in range(2, wb.max_row + 1):
    market = wb[f"C{i}"].value
    item_code = wb[f"D{i}"].value
    operation_type = wb[f"F{i}"].value
    package_count = wb[f"E{i}"].value

    if market in needed_markets and item_code == 23:
        if operation_type == "Поступление" and package_count is not None:
            in_stock += package_count
            #print(f"added {package_count} on row {i}")
        elif operation_type == "Продажа" and package_count is not None:
            in_stock -= package_count
            #print(f"sent {package_count} on row {i}")

print(in_stock)