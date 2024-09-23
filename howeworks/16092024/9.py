from openpyxl import load_workbook
from collections import Counter

def evaluate_row(row_data):
    values = [int(cell.value) for cell in row_data]

    value_counts = Counter(values)
    largest_value = max(values)

    repeated_values = any(freq > 1 for freq in value_counts.values())
    largest_not_repeated = value_counts[largest_value] == 1
    repeated_sum = sum(num * freq for num, freq in value_counts.items() if freq > 1)

    return repeated_values and largest_not_repeated and repeated_sum < largest_value

def process_excel(file_name):
    workbook = load_workbook(file_name)
    worksheet = workbook.active

    valid_rows = 0

    for row in worksheet.iter_rows(min_row=1, max_col=6):
        if evaluate_row(row):
            valid_rows += 1

    return valid_rows

path_to_file = '09.xlsx'
final_result = process_excel(path_to_file)
print(final_result)
