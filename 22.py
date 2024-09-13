import openpyxl

workbook = openpyxl.load_workbook("22.xlsx")
sheet = workbook["Лист1"]

process_data = {}

for row in range(2, sheet.max_row + 1):
    item_id = sheet[f"A{row}"].value
    sender = sheet[f"C{row}"].value
    timestamp = sheet[f"B{row}"].value

    if sender == 0:
        process_data[item_id] = timestamp

print(process_data.keys())

total_count = 0

for row in range(2, sheet.max_row + 1):
    item_id = sheet[f"A{row}"].value
    sender = sheet[f"C{row}"].value
    timestamp = sheet[f"B{row}"].value

    if sender == 0:
        continue
    else:
        total_time = 0
        sender_list = str(sender).split(";")
        print(sender_list)

        for sender_id in sender_list:
            total_time += process_data[int(sender_id)]

        total_time += timestamp

        if total_time >= 100:
            total_count += 1

        print(f"id: {item_id}, time: {total_time}")
        process_data[item_id] = total_time

print(total_count)